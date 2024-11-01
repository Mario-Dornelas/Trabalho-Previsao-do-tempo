from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os
import tkinter as tk
from tkinter import messagebox

def buscar_previsao():
    try:
        
        driver = webdriver.Chrome()  
        driver.get("https://www.google.com/search?q=previsao+sao+paulo")  

        
        temperatura_element = WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.XPATH, '//*[@id="wob_tm"]'))
        )
        umidade_element = WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.XPATH, '//*[@id="wob_hm"]'))
        )

        
        temperatura = temperatura_element.text  
        umidade = umidade_element.text          

        
        print(f"Temperatura: {temperatura}")
        print(f"Umidade: {umidade}")

        
        driver.quit()

        
        salvar_dados(temperatura, umidade)

        
        messagebox.showinfo("Sucesso", "Dados salvos com sucesso!")

    except Exception as e:
        print("Erro ao buscar a previsão:", e)
        driver.quit()
        messagebox.showerror("Erro", "Erro ao buscar a previsão.")

def salvar_dados(temperatura, umidade):
    try:
        
        caminho_arquivo = os.path.join(os.path.dirname(os.path.abspath(__file__)), "temperatura_sao_paulo.xlsx")

        
        print(f"Caminho onde a planilha será salva: {caminho_arquivo}")

        
        try:
            workbook = load_workbook(caminho_arquivo)
            sheet = workbook.active
            print("Arquivo encontrado, abrindo para atualização.")
        except FileNotFoundError:
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["Data/Hora", "Temperatura", "Umidade"])  # Cabeçalhos
            print("Arquivo não encontrado, criando um novo.")

       
        sheet.append([datetime.now().strftime("%Y-%m-%d %H:%M:%S"), temperatura, umidade])
        workbook.save(caminho_arquivo)

        print(f"Dados salvos com sucesso em: {caminho_arquivo}")

    except Exception as e:
        print("Erro ao salvar os dados:", e)


root = tk.Tk()
root.title("Captador de Temperatura de São Paulo")


label = tk.Label(root, text="Atualizar previsão na planilha")
label.pack(pady=10)


btn_buscar = tk.Button(root, text="Buscar Previsão", command=buscar_previsao)
btn_buscar.pack(pady=20)


root.mainloop()
